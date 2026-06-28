"""
04_excel_report.py — Per-vendor Excel report with embedded matplotlib charts.

Reads from: analysis.db  (written by 02_filter.py)

Usage:
    python3 04_excel_report.py --family SDWAN
    python3 04_excel_report.py --all      # generate one file per family

Each report contains:
  - Summary sheet : vendor × year matrix + stacked-by-year bar chart
  - All CVEs sheet: full CVE list across all vendors
  - Per-vendor sheets: year summary table + 3 charts + CVE detail table
"""

import argparse
import io
import re
import sqlite3
from pathlib import Path

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── Palette ───────────────────────────────────────────────────────────────────

CLR_NAVY    = "1F3864"
CLR_BLUE    = "2E75B6"
CLR_WHITE   = "FFFFFF"
CLR_ALT     = "EBF3FB"
CLR_GREEN   = "#70AD47"   # score <5
CLR_AMBER   = "#FFC000"   # score 5-8
CLR_RED     = "#C00000"   # score >8
CLR_TREND   = "#2E75B6"   # line chart

# Severity row highlight colours (light fills)
SEV_FILL = {
    "CRITICAL": "FFD7D7",
    "HIGH":     "FFE8CC",
    "MEDIUM":   "FFFACD",
    "LOW":      "E2EFDA",
}
SEV_BADGE = {           # darker badge for the severity cell itself
    "CRITICAL": "C00000",
    "HIGH":     "ED7D31",
    "MEDIUM":   "FFC000",
    "LOW":      "70AD47",
}

CVE_COLS = [            # (header, width, align)
    ("CVE ID",       18, "left"),
    ("Published",    13, "center"),
    ("CVSS Score",   11, "center"),
    ("Severity",     11, "center"),
    ("CWE(s)",       16, "center"),
    ("Vendors",      22, "left"),
    ("Description", 100, "left"),
]

SCORE_BUCKETS = ["<5", "5-8", ">8"]
SCORE_COLORS  = [CLR_GREEN, CLR_AMBER, CLR_RED]
SCORE_LABELS  = ["Low (<5)", "Medium (5-8)", "High (>8)"]

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.grid": True,
    "grid.alpha": 0.3,
    "axes.titlesize": 13,
    "axes.titleweight": "bold",
    "axes.labelsize": 10,
    "xtick.labelsize": 9,
    "ytick.labelsize": 9,
})

# ── Cell helpers ──────────────────────────────────────────────────────────────

def hcell(ws, row, col, val, bg=CLR_NAVY, fg=CLR_WHITE, size=11, bold=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, size=size, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c

def dcell(ws, row, col, val, bold=False, align="center", bg=None, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

# ── Score bucket ─────────────────────────────────────────────────────────────

def score_bucket(score):
    if score is None or (isinstance(score, float) and np.isnan(score)):
        return "<5"
    if score < 5:
        return "<5"
    elif score <= 8:
        return "5-8"
    return ">8"

# ── Data loading ─────────────────────────────────────────────────────────────

def load_data(db_path: str, family: str) -> tuple[dict[str, pd.DataFrame], list[int]]:
    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query(
        "SELECT * FROM filtered_cves WHERE family = ?", conn, params=(family,)
    )
    conn.close()

    if df.empty:
        raise SystemExit(f"[ERROR] No CVEs for family '{family}' — run 02_filter.py first.")

    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.dropna(subset=["year"])
    df["year"] = df["year"].astype(int)
    df["cvss_score"] = pd.to_numeric(df["cvss_score"], errors="coerce")
    df["score_bucket"] = df["cvss_score"].apply(score_bucket)

    # Explode vendors — drop "Generic"; if only Generic, skip CVE entirely
    rows = []
    for _, r in df.iterrows():
        vendors = [v.strip() for v in str(r.get("vendors", "")).split(",") if v.strip()]
        specific = [v for v in vendors if v != "Generic"]
        for v in specific:
            rows.append({**r.to_dict(), "vendor": v})

    vdf = pd.DataFrame(rows)
    if vdf.empty:
        raise SystemExit("[ERROR] No vendor-specific CVEs found after dropping Generic.")

    all_years = sorted(vdf["year"].unique().tolist())

    vendor_map = {}
    for vendor, grp in vdf.groupby("vendor"):
        vendor_map[vendor] = grp.copy()

    return vendor_map, all_years


def build_pivot(grp: pd.DataFrame, all_years: list[int]) -> pd.DataFrame:
    p = (
        grp.groupby(["year", "score_bucket"])
        .size()
        .unstack(fill_value=0)
        .reindex(columns=SCORE_BUCKETS, fill_value=0)
    )
    p["total"] = p[SCORE_BUCKETS].sum(axis=1)
    p = p.reindex(all_years, fill_value=0)
    p.index.name = "year"
    return p.reset_index()

# ── Chart generators ──────────────────────────────────────────────────────────

FIG_W, FIG_H = 9, 4.5   # inches


def fig_to_img(fig) -> io.BytesIO:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor="white")
    buf.seek(0)
    plt.close(fig)
    return buf


def chart_bar_total(pivot: pd.DataFrame, vendor: str) -> io.BytesIO:
    years = pivot["year"].astype(str).tolist()
    totals = pivot["total"].tolist()

    fig, ax = plt.subplots(figsize=(FIG_W, FIG_H))
    bars = ax.bar(years, totals, color=CLR_TREND, edgecolor="white", width=0.6, zorder=3)

    # Count labels on top of each bar
    for bar, val in zip(bars, totals):
        if val > 0:
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + max(totals) * 0.015,
                str(int(val)),
                ha="center", va="bottom", fontsize=9, fontweight="bold", color="#333333"
            )

    ax.set_title(f"{vendor} — CVE Count per Year", pad=12)
    ax.set_xlabel("Year")
    ax.set_ylabel("Number of CVEs")
    ax.set_ylim(0, max(totals) * 1.18 if max(totals) > 0 else 10)
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    return fig_to_img(fig)


def chart_stacked_scores(pivot: pd.DataFrame, vendor: str) -> io.BytesIO:
    years = pivot["year"].astype(str).tolist()
    x = np.arange(len(years))
    width = 0.6

    fig, ax = plt.subplots(figsize=(FIG_W, FIG_H))
    bottoms = np.zeros(len(years))

    for bucket, color, label in zip(SCORE_BUCKETS, SCORE_COLORS, SCORE_LABELS):
        vals = pivot[bucket].values.astype(float)
        bars = ax.bar(x, vals, width, bottom=bottoms, color=color,
                      label=label, edgecolor="white", zorder=3)
        # Segment labels (only if segment tall enough)
        for i, (bar, v) in enumerate(zip(bars, vals)):
            if v >= 1:
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bottoms[i] + v / 2,
                    str(int(v)),
                    ha="center", va="center", fontsize=8,
                    fontweight="bold", color="white"
                )
        bottoms += vals

    # Total label on top of each stack
    totals = pivot["total"].values
    for i, (tot, bot) in enumerate(zip(totals, bottoms)):
        if tot > 0:
            ax.text(i, bot + max(totals) * 0.01, str(int(tot)),
                    ha="center", va="bottom", fontsize=9, fontweight="bold", color="#333333")

    ax.set_title(f"{vendor} — CVE Score Buckets per Year", pad=12)
    ax.set_xlabel("Year")
    ax.set_ylabel("Number of CVEs")
    ax.set_xticks(x)
    ax.set_xticklabels(years, rotation=45, ha="right")
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax.set_ylim(0, max(totals) * 1.18 if max(totals) > 0 else 10)
    ax.legend(loc="upper left", framealpha=0.85, fontsize=9)
    plt.tight_layout()
    return fig_to_img(fig)


def chart_trend_line(pivot: pd.DataFrame, vendor: str) -> io.BytesIO:
    years = pivot["year"].astype(str).tolist()
    totals = pivot["total"].values.astype(float)

    fig, ax = plt.subplots(figsize=(FIG_W, FIG_H))
    ax.fill_between(range(len(years)), totals, alpha=0.15, color=CLR_TREND)
    ax.plot(range(len(years)), totals, color=CLR_TREND, linewidth=2.5,
            marker="o", markersize=7, markerfacecolor="white",
            markeredgecolor=CLR_TREND, markeredgewidth=2, zorder=4)

    # Point labels
    for i, (v, tot) in enumerate(zip(range(len(years)), totals)):
        if tot >= 0:
            ax.text(i, tot + max(totals) * 0.04 + 0.1, str(int(tot)),
                    ha="center", va="bottom", fontsize=9,
                    fontweight="bold", color="#333333")

    ax.set_title(f"{vendor} — CVEs Over Time (Trend)", pad=12)
    ax.set_xlabel("Year")
    ax.set_ylabel("Number of CVEs")
    ax.set_xticks(range(len(years)))
    ax.set_xticklabels(years, rotation=45, ha="right")
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax.set_ylim(0, max(totals) * 1.22 if max(totals) > 0 else 10)
    plt.tight_layout()
    return fig_to_img(fig)


def chart_summary_stacked_by_year(all_pivots: dict[str, pd.DataFrame], all_years: list[int]) -> io.BytesIO:
    """Stacked bar chart: X-axis = year, each colour segment = vendor."""
    vendors = sorted(all_pivots, key=lambda v: all_pivots[v]["total"].sum(), reverse=True)

    # One distinct colour per vendor
    cmap = matplotlib.colormaps["tab10"].resampled(max(len(vendors), 10))
    vendor_colours = {v: cmap(i) for i, v in enumerate(vendors)}

    x       = np.arange(len(all_years))
    bottoms = np.zeros(len(all_years))
    fig, ax = plt.subplots(figsize=(max(12, len(all_years) * 0.95), 8))

    for vendor in vendors:
        pivot = all_pivots[vendor].set_index("year")
        vals  = np.array([
            int(pivot.loc[yr, "total"]) if yr in pivot.index else 0
            for yr in all_years
        ], dtype=float)
        bars = ax.bar(x, vals, bottom=bottoms, width=0.65,
                      color=vendor_colours[vendor], label=vendor,
                      edgecolor="white", linewidth=0.6, zorder=3)

        # Segment labels inside bar (only if segment is tall enough)
        for i, (bar, v) in enumerate(zip(bars, vals)):
            if v >= 3:
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bottoms[i] + v / 2,
                        str(int(v)),
                        ha="center", va="center",
                        fontsize=7.5, fontweight="bold", color="white")
        bottoms += vals

    # Total labels above each bar
    max_total = bottoms.max() if bottoms.max() > 0 else 1
    for i, total in enumerate(bottoms):
        if total > 0:
            ax.text(i, total + max_total * 0.012, str(int(total)),
                    ha="center", va="bottom", fontsize=10,
                    fontweight="bold", color="#444444")

    ax.set_title(f"SD-WAN CVEs by Year  ({all_years[0]}–{all_years[-1]})",
                 fontsize=15, fontweight="bold", pad=16)
    ax.set_ylabel("Number of CVEs", fontsize=11)
    ax.set_xticks(x)
    ax.set_xticklabels([str(y) for y in all_years], fontsize=10)
    ax.set_ylim(0, max_total * 1.14)
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", alpha=0.25, zorder=0)

    ax.legend(title="Vendor", title_fontsize=10,
              loc="upper left", bbox_to_anchor=(1.01, 1),
              framealpha=0.9, fontsize=9, borderaxespad=0)

    plt.tight_layout()
    return fig_to_img(fig)

# ── CVE detail table ─────────────────────────────────────────────────────────

def write_cve_table(ws, cve_df: pd.DataFrame, start_row: int, title: str):
    """Write a sortable CVE detail table starting at start_row."""

    # Section title
    last_col = len(CVE_COLS)
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=last_col
    )
    t = ws.cell(row=start_row, column=1, value=title)
    t.font = Font(name="Arial", bold=True, size=12, color=CLR_WHITE)
    t.fill = PatternFill("solid", fgColor=CLR_NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 20

    # Column headers
    hdr_row = start_row + 1
    for col, (hdr, width, _) in enumerate(CVE_COLS, 1):
        hcell(ws, hdr_row, col, hdr, bg=CLR_BLUE)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[hdr_row].height = 22

    # Sort: critical first, then by score desc, then date desc
    sev_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    df = cve_df.copy()
    df["_sev_rank"] = df["cvss_severity"].str.upper().map(sev_order).fillna(4)
    df["cvss_score"] = pd.to_numeric(df["cvss_score"], errors="coerce")
    df = df.sort_values(["_sev_rank", "cvss_score", "published"],
                        ascending=[True, False, False])

    for i, (_, row) in enumerate(df.iterrows()):
        dr = hdr_row + 1 + i
        sev = str(row.get("cvss_severity") or "").upper()
        row_bg  = SEV_FILL.get(sev)
        score   = row.get("cvss_score")
        pub     = str(row.get("published") or "")[:10]   # YYYY-MM-DD
        desc    = str(row.get("description") or "")
        desc_short = (desc[:300] + "…") if len(desc) > 300 else desc
        cve_id  = str(row.get("cve_id") or "")
        cwes    = str(row.get("weaknesses") or "").replace(",", ", ")
        vendors = str(row.get("vendors") or "")

        # CVE ID — hyperlink to NVD
        c = ws.cell(row=dr, column=1, value=cve_id)
        c.font = Font(name="Arial", size=9, color="0563C1", underline="single")
        c.hyperlink = f"https://nvd.nist.gov/vuln/detail/{cve_id}"
        c.alignment = Alignment(horizontal="left", vertical="top")
        c.border = BORDER
        if row_bg:
            c.fill = PatternFill("solid", fgColor=row_bg)

        # Published
        _dc(ws, dr, 2, pub, bg=row_bg, align="center")

        # CVSS Score
        score_val = float(score) if pd.notna(score) else None
        sc = _dc(ws, dr, 3, score_val, bg=row_bg, align="center", fmt="0.0")
        if score_val is not None:
            sc.font = Font(name="Arial", size=9, bold=True)

        # Severity badge
        badge_bg = SEV_BADGE.get(sev, "BFBFBF")
        sv = ws.cell(row=dr, column=4, value=sev if sev else "N/A")
        sv.font = Font(name="Arial", size=9, bold=True, color=CLR_WHITE)
        sv.fill = PatternFill("solid", fgColor=badge_bg)
        sv.alignment = Alignment(horizontal="center", vertical="top")
        sv.border = BORDER

        # CWEs
        _dc(ws, dr, 5, cwes, bg=row_bg, align="center")

        # Vendors
        _dc(ws, dr, 6, vendors, bg=row_bg, align="left")

        # Description
        dc = _dc(ws, dr, 7, desc_short, bg=row_bg, align="left")
        dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        dc.font = Font(name="Arial", size=9)

        ws.row_dimensions[dr].height = 42

    return hdr_row + 1 + len(df)   # return next free row


def _dc(ws, row, col, val, bg=None, align="center", fmt=None):
    """Compact data cell for CVE table."""
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=9)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=False)
    c.border = BORDER
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c


# ── Sheet writers ─────────────────────────────────────────────────────────────

IMG_ROW_START = 4   # data table starts row 4
IMG_COL       = 7   # G — charts go here

# Approximate row heights to position charts (each chart ~22 rows tall at default zoom)
CHART_ROW_HEIGHT = 22


def insert_img(ws, buf: io.BytesIO, anchor: str, w_cm=17, h_cm=9):
    img = XLImage(buf)
    img.width  = int(w_cm * 37.795)   # cm → pixels (approx)
    img.height = int(h_cm * 37.795)
    img.anchor = anchor
    ws.add_image(img)


def write_vendor_sheet(wb: Workbook, vendor: str, pivot: pd.DataFrame, cve_df: pd.DataFrame):
    safe = re.sub(r'[\\/*?\[\]:]', '-', vendor)[:31]
    ws = wb.create_sheet(title=safe)

    n_rows = len(pivot)
    hdr_row   = 3
    data_row0 = 4
    data_rowN = data_row0 + n_rows - 1
    total_row = data_rowN + 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:E2")
    t = ws["A1"]
    t.value = f"{vendor}  |  SD-WAN CVE Analysis"
    t.font = Font(name="Arial", bold=True, size=14, color=CLR_WHITE)
    t.fill = PatternFill("solid", fgColor=CLR_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    # ── Column headers ────────────────────────────────────────────────────────
    hcell(ws, hdr_row, 1, "Year",          bg=CLR_BLUE)
    hcell(ws, hdr_row, 2, "Total CVEs",    bg=CLR_BLUE)
    hcell(ws, hdr_row, 3, "Score <5\n(Low)",   bg="375623", fg=CLR_WHITE)
    hcell(ws, hdr_row, 4, "Score 5-8\n(Med)",  bg="7F6000", fg=CLR_WHITE)
    hcell(ws, hdr_row, 5, "Score >8\n(High)",  bg="7B0000", fg=CLR_WHITE)
    ws.row_dimensions[hdr_row].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, row in pivot.iterrows():
        dr = data_row0 + i
        bg = CLR_ALT if i % 2 == 0 else None
        dcell(ws, dr, 1, int(row["year"]),  bold=True, bg=bg)
        dcell(ws, dr, 2, int(row["total"]), bg=bg, fmt="#,##0")
        dcell(ws, dr, 3, int(row["<5"]),    bg=bg, fmt="#,##0")
        dcell(ws, dr, 4, int(row["5-8"]),   bg=bg, fmt="#,##0")
        dcell(ws, dr, 5, int(row[">8"]),    bg=bg, fmt="#,##0")

    # ── Totals row ────────────────────────────────────────────────────────────
    hcell(ws, total_row, 1, "TOTAL", bg=CLR_BLUE)
    for col in range(2, 6):
        cl = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUM({cl}{data_row0}:{cl}{data_rowN})")
        c.font = Font(name="Arial", bold=True, color=CLR_WHITE)
        c.fill = PatternFill("solid", fgColor=CLR_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        c.number_format = "#,##0"

    # Column widths
    for col, w in zip("ABCDE", [9, 12, 13, 13, 13]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["F"].width = 2

    # ── Charts ────────────────────────────────────────────────────────────────
    chart_w_cm, chart_h_cm = 17, 9

    img1 = chart_bar_total(pivot, vendor)
    insert_img(ws, img1, f"G{hdr_row}", chart_w_cm, chart_h_cm)

    img2 = chart_stacked_scores(pivot, vendor)
    anchor2_row = hdr_row + CHART_ROW_HEIGHT
    insert_img(ws, img2, f"G{anchor2_row}", chart_w_cm, chart_h_cm)

    img3 = chart_trend_line(pivot, vendor)
    anchor3_row = anchor2_row + CHART_ROW_HEIGHT
    insert_img(ws, img3, f"G{anchor3_row}", chart_w_cm, chart_h_cm)

    # ── CVE Detail Table (below charts) ───────────────────────────────────────
    cve_table_row = anchor3_row + CHART_ROW_HEIGHT + 2
    write_cve_table(ws, cve_df, cve_table_row,
                    f"{vendor} — Individual CVEs ({len(cve_df)} total, sorted by severity)")


def write_all_cves_sheet(wb: Workbook, vendor_data: dict[str, pd.DataFrame]):
    """Master sheet listing every CVE across all vendors."""
    ws = wb.create_sheet(title="All CVEs")

    # Combine all vendor frames, keep one row per unique CVE (deduplicate)
    combined = pd.concat(vendor_data.values(), ignore_index=True)
    combined = combined.drop_duplicates(subset="cve_id")
    total = len(combined)

    write_cve_table(ws, combined, 1,
                    f"All SD-WAN CVEs — {total} unique entries across all vendors")


def write_summary_sheet(wb: Workbook, all_pivots: dict[str, pd.DataFrame], all_years: list[int]):
    ws = wb.active
    ws.title = "Summary"

    vendors = sorted(all_pivots, key=lambda v: all_pivots[v]["total"].sum(), reverse=True)
    n_year_cols = len(all_years)
    last_col = get_column_letter(3 + n_year_cols - 1)

    ws.merge_cells(f"A1:{last_col}2")
    t = ws["A1"]
    t.value = "SD-WAN CVE Trend Analysis  |  All Vendors"
    t.font = Font(name="Arial", bold=True, size=15, color=CLR_WHITE)
    t.fill = PatternFill("solid", fgColor=CLR_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # Header
    r = 3
    hcell(ws, r, 1, "Vendor",  bg=CLR_BLUE)
    hcell(ws, r, 2, "Total",   bg=CLR_BLUE)
    for j, yr in enumerate(all_years, 3):
        hcell(ws, r, j, str(yr), bg=CLR_BLUE)
    ws.row_dimensions[r].height = 22

    # Rows
    for i, vendor in enumerate(vendors):
        dr = r + 1 + i
        bg = CLR_ALT if i % 2 == 0 else None
        pivot = all_pivots[vendor].set_index("year")
        total = int(pivot["total"].sum())
        dcell(ws, dr, 1, vendor, align="left", bg=bg, bold=True)
        dcell(ws, dr, 2, total,  bg=bg, fmt="#,##0")
        for j, yr in enumerate(all_years, 3):
            val = int(pivot.loc[yr, "total"]) if yr in pivot.index else 0
            dcell(ws, dr, j, val, bg=bg, fmt="#,##0")

    # Totals row
    tr = r + 1 + len(vendors)
    hcell(ws, tr, 1, "TOTAL", bg=CLR_BLUE)
    for col in range(2, 3 + n_year_cols):
        cl = get_column_letter(col)
        c = ws.cell(row=tr, column=col,
                    value=f"=SUM({cl}{r+1}:{cl}{tr-1})")
        c.font = Font(name="Arial", bold=True, color=CLR_WHITE)
        c.fill = PatternFill("solid", fgColor=CLR_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        c.number_format = "#,##0"

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 10
    for j in range(n_year_cols):
        ws.column_dimensions[get_column_letter(3 + j)].width = 7

    # Summary chart — stacked by year, one bar per vendor
    img = chart_summary_stacked_by_year(all_pivots, all_years)
    anchor_row = tr + 3
    insert_img(ws, img, f"A{anchor_row}", w_cm=30, h_cm=14)


# ── Main ─────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--db",  default="analysis.db",    help="Source DB (from 02_filter.py)")
    p.add_argument("--out", default="sdwan_report.xlsx", help="Output Excel file")
    return p.parse_args()


def main():
    args = parse_args()
    print(f"[Excel] Loading data from {args.db}...")
    vendor_data, all_years = load_data(args.db, "SDWAN")
    print(f"[Excel] Years {all_years[0]}–{all_years[-1]} | Vendors: {sorted(vendor_data)}")

    all_pivots = {v: build_pivot(grp, all_years) for v, grp in vendor_data.items()}

    wb = Workbook()
    write_summary_sheet(wb, all_pivots, all_years)
    write_all_cves_sheet(wb, vendor_data)

    for vendor in sorted(all_pivots, key=lambda v: all_pivots[v]["total"].sum(), reverse=True):
        print(f"  → {vendor}")
        write_vendor_sheet(wb, vendor, all_pivots[vendor], vendor_data[vendor])

    wb.save(Path(args.out))
    print(f"\n[Done] Saved → {args.out}")


if __name__ == "__main__":
    main()
